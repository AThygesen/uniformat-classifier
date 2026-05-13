"""Excel I/O for tenderlist files.

All historical and template tenderlist files share the same shape:
- Single data sheet (usually 'Sheet1', or 'Template' in the template file)
- Row 1 (Excel) = column headers
- Row 2 (Excel) = a marker row like "tekst" / "Do not delete this line"
- Row 3+ (Excel) = data

Key columns we read/write by header name (not column index, since some files vary slightly):
- 'Ydelse'                           — the description (model input)
- 'Niveau 1', 'Niveau 2', 'Niveau 3' — the labels we predict
- 'Benchmarking klassifikation'      — also written
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

YDELSE_COL = "Ydelse"
NIVEAU_1_COL = "Niveau 1"
NIVEAU_2_COL = "Niveau 2"
NIVEAU_3_COL = "Niveau 3"
BENCHMARKING_COL = "Benchmarking klassifikation"

# Headers we look for to find the data sheet (some files use 'Sheet1', the template uses 'Template')
HEADER_SENTINELS = (YDELSE_COL,)


@dataclass
class TenderRow:
    """A single tenderlist row, normalized for the classifier."""
    sheet_row_ix: int           # 1-indexed Excel row number (for write-back)
    ydelse: str                 # the description text
    niveau1: str | None
    niveau2: str | None
    niveau3: str | None
    benchmarking: str | None

    @property
    def is_labeled(self) -> bool:
        """A row counts as labeled if it has at least Niveau 1 (the most-common case in the data)."""
        return bool(self.niveau1)

    @property
    def has_niveau3(self) -> bool:
        return bool(self.niveau3)


def find_data_sheet(wb: Workbook) -> Worksheet:
    """Return the worksheet whose first row contains 'Ydelse'."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # iter the first row only
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first and YDELSE_COL in [str(c).strip() if c is not None else "" for c in first]:
            return ws
    raise ValueError(
        f"No sheet with column '{YDELSE_COL}' found. Sheets: {wb.sheetnames}"
    )


def header_index(ws: Worksheet) -> dict[str, int]:
    """Return {header_name: 1-indexed column number} from the first row."""
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    out: dict[str, int] = {}
    for cell in first:
        v = cell.value
        if v is None:
            continue
        out[str(v).strip()] = cell.column
    return out


def read_rows(path: Path) -> list[TenderRow]:
    """Read all data rows from a tenderlist Excel file."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = find_data_sheet(wb)
        idx = header_index(ws)

        col_yd = idx.get(YDELSE_COL)
        col_n1 = idx.get(NIVEAU_1_COL)
        col_n2 = idx.get(NIVEAU_2_COL)
        col_n3 = idx.get(NIVEAU_3_COL)
        col_bm = idx.get(BENCHMARKING_COL)

        if col_yd is None:
            raise ValueError(f"{path.name}: missing required column '{YDELSE_COL}'")

        rows: list[TenderRow] = []
        # Skip header row (1) and the marker row (2). Data starts at row 3.
        for r_ix, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            ydelse = _cell(row, col_yd)
            if not ydelse or not str(ydelse).strip():
                continue  # skip empty/footer rows
            rows.append(
                TenderRow(
                    sheet_row_ix=r_ix,
                    ydelse=str(ydelse).strip(),
                    niveau1=_str_or_none(_cell(row, col_n1)),
                    niveau2=_str_or_none(_cell(row, col_n2)),
                    niveau3=_str_or_none(_cell(row, col_n3)),
                    benchmarking=_str_or_none(_cell(row, col_bm)),
                )
            )
        return rows
    finally:
        wb.close()


def write_back(
    path: Path,
    updates: dict[int, dict[str, str]],
    *,
    backup: bool = True,
) -> None:
    """Write classification updates back into an Excel file in-place.

    Args:
        path: file to update.
        updates: {sheet_row_ix: {header_name: value, ...}} — only the given headers are written.
        backup: if True, copy original to <name>.bak.xlsx before writing.
    """
    if backup:
        bak = path.with_suffix(path.suffix + ".bak")
        if not bak.exists():
            bak.write_bytes(path.read_bytes())

    wb = load_workbook(path)  # read/write mode
    try:
        ws = find_data_sheet(wb)
        idx = header_index(ws)
        for r_ix, fields in updates.items():
            for header, value in fields.items():
                col = idx.get(header)
                if col is None:
                    # Header not present in this file — skip silently rather than crash.
                    continue
                ws.cell(row=r_ix, column=col, value=value)
        wb.save(path)
    finally:
        wb.close()


def _cell(row: tuple, col_1indexed: int | None):
    if col_1indexed is None:
        return None
    ix = col_1indexed - 1
    if ix < 0 or ix >= len(row):
        return None
    return row[ix]


def _str_or_none(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
