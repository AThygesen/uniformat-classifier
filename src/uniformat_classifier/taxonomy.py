"""Uniformat taxonomy loaded from Indskrivning_template.xlsx → Central_classifikation."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

CENTRAL_SHEET = "Central_classifikation"

# Header row layout (0-indexed within Central_classifikation):
# row 2 is the header, row 3+ is data.
# Columns (1-indexed in the sheet, 0-indexed when iter_rows returns tuples):
#   col B (1): Position (running index)
#   col C (2): Uniformat Kode    (e.g. "A", "A10", "A1010", "A101001")
#   col D (3): Sammen            (e.g. "A1010. Standard Foundations - Standard Fundamenter")
#   col E (4): Beskrivelse       (just the text portion)
#   col F (5): Benchmarking klassifikation
HEADER_ROW_IX = 2
DATA_START_IX = 3
COL_CODE = 2
COL_SAMMEN = 3
COL_BESKRIVELSE = 4
COL_BENCHMARKING = 5


@dataclass(frozen=True)
class TaxonomyEntry:
    code: str            # e.g. "A1010"
    label: str           # full "Sammen" string, e.g. "A1010. Standard Foundations - Standard Fundamenter"
    description: str     # just the descriptive part
    benchmarking: str    # Benchmarking klassifikation (e.g. "A. Substructure - Bygningsbasis")
    level: int           # 1=A, 2=A10, 3=A1010, 4=A101001 (Danish-specific extension)


class Taxonomy:
    """In-memory Uniformat taxonomy with rollup helpers."""

    def __init__(self, entries: list[TaxonomyEntry]) -> None:
        self._by_code: dict[str, TaxonomyEntry] = {e.code: e for e in entries}
        self._entries = entries

    # ----- factory ----------------------------------------------------------

    @classmethod
    def from_template(cls, template_path: Path) -> "Taxonomy":
        wb = load_workbook(template_path, read_only=True, data_only=True)
        try:
            ws = wb[CENTRAL_SHEET]
        except KeyError as exc:
            raise ValueError(
                f"Sheet '{CENTRAL_SHEET}' not found in {template_path}. "
                f"Available: {wb.sheetnames}"
            ) from exc

        entries: list[TaxonomyEntry] = []
        for r in ws.iter_rows(min_row=DATA_START_IX + 1, values_only=True):
            code = r[COL_CODE] if len(r) > COL_CODE else None
            if not code or not isinstance(code, str):
                continue
            sammen = r[COL_SAMMEN] if len(r) > COL_SAMMEN else None
            beskr = r[COL_BESKRIVELSE] if len(r) > COL_BESKRIVELSE else None
            bench = r[COL_BENCHMARKING] if len(r) > COL_BENCHMARKING else None
            entries.append(
                TaxonomyEntry(
                    code=code.strip(),
                    label=(sammen or code).strip(),
                    description=(beskr or "").strip() if beskr else "",
                    benchmarking=(bench or "").strip() if bench else "",
                    level=_level_from_code(code.strip()),
                )
            )
        wb.close()
        return cls(entries)

    # ----- lookups ----------------------------------------------------------

    def get(self, code: str) -> TaxonomyEntry | None:
        return self._by_code.get(code)

    def all(self) -> list[TaxonomyEntry]:
        return list(self._entries)

    def at_level(self, level: int) -> list[TaxonomyEntry]:
        return [e for e in self._entries if e.level == level]

    # Niveau 3 (5-char) is our prediction target.
    def niveau3(self) -> list[TaxonomyEntry]:
        return self.at_level(3)

    # ----- rollup helpers ---------------------------------------------------

    def parent_code(self, code: str) -> str | None:
        """Return the parent code one level up (A1010 -> A10, A10 -> A, A -> None)."""
        if len(code) >= 7:
            return code[:5]
        if len(code) == 5:
            return code[:3]
        if len(code) == 3:
            return code[:1]
        return None

    def rollup(self, code: str) -> dict[int, TaxonomyEntry]:
        """Return {level: entry} for the code and all its ancestors."""
        out: dict[int, TaxonomyEntry] = {}
        cur: str | None = code
        while cur is not None:
            entry = self._by_code.get(cur)
            if entry is not None:
                out[entry.level] = entry
            cur = self.parent_code(cur)
        return out

    def labels_for(self, code: str) -> tuple[str | None, str | None, str | None]:
        """Return (Niveau 1 label, Niveau 2 label, Niveau 3 label) for any code."""
        roll = self.rollup(code)
        return (
            roll[1].label if 1 in roll else None,
            roll[2].label if 2 in roll else None,
            roll[3].label if 3 in roll else None,
        )

    # ----- reverse lookup from labels seen in legacy files ------------------

    def code_from_label(self, label: str) -> str | None:
        """Match a 'Sammen' label like 'A1010. Standard Foundations -...' back to its code.

        Useful when ingesting historic Excel files where the columns hold the full label.
        """
        if not label or not isinstance(label, str):
            return None
        label = label.strip()
        # Most labels start with the code followed by ". ".
        if "." in label:
            head = label.split(".", 1)[0].strip()
            if head in self._by_code:
                return head
        # Fallback: scan
        for e in self._entries:
            if e.label.strip() == label:
                return e.code
        return None


def _level_from_code(code: str) -> int:
    n = len(code)
    if n == 1:
        return 1
    if n == 3:
        return 2
    if n == 5:
        return 3
    if n >= 7:
        return 4
    return 0
